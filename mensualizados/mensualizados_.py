# Toma un archivo, lo convierte a un dataFrame, elimina las ultimas 3 columnas 
# Si el archivo viene del area de SALUD PUBLICA y AMBIENTE Y ESPACIO PUBLICO subdividir en oficinas
# Y que cada oficina sea un archivo por separado

import pandas as pd
import numpy as np
import streamlit as st
import io
from openpyxl import load_workbook
from openpyxl.styles import numbers

#---------- Funciones principales -----------

def borrar_ultimas_columnas(df: pd.DataFrame, n: int) -> pd.DataFrame:
    """
    Borra las últimas n columnas de un dataFrame
    
    :param df: DataFrame completo
    :param n: Cantidad de últimas columnas a borrar
    :return: Devuelve el dataFrame original sin las últimas n columnas.
    """

    cant_columnas = df.shape[1]

    columnas_a_eliminar = range(cant_columnas - n, cant_columnas)

    df = df.drop(df.columns[columnas_a_eliminar], axis=1)

    return df


def dividir_oficinas(df: pd.DataFrame) -> list[pd.DataFrame]:
    """
    Precondicion: que el DataFrame pertenezca al área de salud pública/ambiente y espacio público
    
    :param df: Description
    :type df: pd.DataFrame
    :return: Description
    :rtype: list[DataFrame]
    """

    oficinas_unicas = df["Oficina"].unique()
    df_oficinas = []

    for oficina in oficinas_unicas:

        df_oficina = df[df["Oficina"] == oficina]

        df_oficinas.append(df_oficina)

    return df_oficinas


st.title("📝Mensualizados")

st.divider()


opciones = [
    "",
    "AMBIENTE Y ESPACIO PUBLICO",
    "ARSI",
    "CAPITAL HUMANO",
    "DESARROLLO HUMANO Y DEPORTES",
    "EDUCACION, CULTURA Y TRABAJO",
    "GENERAL",
    "GOBIERNO",
    "H.C.D.",
    "HACIENDA Y FINANZAS",
    "JEFATURA DE GABINETE",
    "LEGAL Y TECNICA",
    "PLANEAMIENTO URBANO",
    "PRIVADA",
    "SALUD PUBLICA",
    "SEGURIDAD"
]



st.subheader("Elegir el área del cuál se está subiendo el archivo:")

opcion = st.selectbox(
    "Elegir una opción",
    opciones
)

if opcion == "":
    st.subheader("IMPORTANTE❗: seleccionar el área antes de continuar")

else:

    st.subheader(f"📂Archivo de mensualizados del área {opcion}")

    st.markdown("Subir el archivo de mensualizados")

    archivo_1 = st.file_uploader("Seleccionar el archivo de mensualizados", type=["xlsx", "xls"], key="archivo1",accept_multiple_files=False)

    if archivo_1:

        df = pd.read_excel(archivo_1)
        # 1. Filtrar filas donde la columna no esté vacía
        df = df[df["Fecha Egreso Cargo"].notna() & (df["Fecha Egreso Cargo"] != "")]
        
        # 2. Reemplazar en Categoría
        df["Categoría"] = df["Categoría"].replace("NO CATEGORIZADO", 999)

        if opcion == "SALUD PUBLICA" or opcion == "AMBIENTE Y ESPACIO PUBLICO":
            
            df = borrar_ultimas_columnas(df, 3)

            df_oficinas = dividir_oficinas(df)

            for df_oficina in df_oficinas:
                oficina = df_oficina["Oficina"].unique() # Array de valores unicos
                
                outputi = io.BytesIO()

                df_oficina.to_excel(outputi, index=False)

                outputi.seek(0)
                nombre_archivo_i = f"{opcion}_oficina_{oficina[0]}.xlsx"

                #Para cambiar el formato de la fecha
                wb = load_workbook(outputi)
                ws = wb.active

                columnas_fecha = ["H", "I"] #Columnas en formato fecha

                for col in columnas_fecha:
                    for cell in ws[col]:
                        cell.number_format = "DD/MM/YYYY" #Formato día, mes, año (en el excel se sigue manteniendo tipo de dato fecha)

                outputi2 = io.BytesIO()
                wb.save(outputi2)
                outputi2.seek(0)

                st.download_button(
                    label=f"Descargar planilla de la oficina: {oficina[0]}",
                    data=outputi2.getvalue(),
                    file_name=nombre_archivo_i,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        else:

            df = borrar_ultimas_columnas(df,3)
            outputi = io.BytesIO()

            df.to_excel(outputi, index=False)

            outputi.seek(0)
            nombre_archivo_i = f"{opcion}.xlsx"

            #Para cambiar el formato de la fecha
            wb = load_workbook(outputi)
            ws = wb.active

            columnas_fecha = ["H", "I"] #Columnas en formato fecha

            for col in columnas_fecha:
                for cell in ws[col]:
                    cell.number_format = "DD/MM/YYYY" #Formato en día, mes, año (en el excel sigue manteniendo tipo date)

            outputi2 = io.BytesIO()
            wb.save(outputi2)
            outputi2.seek(0)

            st.download_button(
                label="📂 Descargar planilla de mensualizados",
                data=outputi2.getvalue(),
                file_name=nombre_archivo_i,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


